from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Side, Font
from openpyxl.styles.borders import Border
import re
import glob

files = glob.glob('*.xlsx')  # grabs all spreadsheets in the folder and puts them in a list

for file in files:  # loops through all the spreadsheets in list and formats them
    wb = load_workbook(str(file))  # opens the workbook
    ws = wb.active  # loads the active sheet
    reportFont = Font(size=18, bold=True, italic=False)
    reportNameCell = ws['B4']
    reportNameCell.font = reportFont
    reportName = ws['B4'].value  # Gets job id and name from report
    reportName = re.sub('\\\\', '', reportName)
    reportName = re.sub('/', '', reportName)
    header = ws['A7']  # defines section to freeze
    ws.freeze_panes = header  # freezer top 7 rows
    costFill = PatternFill(start_color='e0fdff',  # color for cost
                           end_color='e0fdff',
                           fill_type='solid')
    whiteFill = PatternFill(start_color='FFFFFFFF',  # color to format all cells white
                            end_color='FFFFFFFF',
                            fill_type='solid')
    totalsFill = PatternFill(start_color='f9fcae',  # color to totals0 cells yellow
                             end_color='f9fcae',
                             fill_type='solid')
    thin_border = Border(left=Side(style='hair'),  # Puts thin border around cells because whitefill erases default ones
                         right=Side(style='hair'),
                         top=Side(style='hair'),
                         bottom=Side(style='hair'))
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    # Puts thin border around cells because whitefill erases default ones
    # thin_border = Border(left=Side(style='hair'),
    #                      right=Side(style='hair'),
    #                      top=Side(style='hair'),
    #                      bottom=Side(style='hair'))

    maxRow = ws.max_row  # defines the last row
    maxCol = ws.max_column  # defines the last column
    finish = "18 - Finish"  # looks for the finish section step 1
    sub = "Sub Totals:"  # defines the sub total so we can find the row
    ws.column_dimensions['C'].width = 20  # makes column C wider
    ws.column_dimensions['E'].hidden = True  # hides column E
    ws.column_dimensions['H'].width = 9  # makes column H wider
    ws.column_dimensions['I'].width = 15  # makes column I wider

    for col in ['D', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    # Formats the Header
    for formCol in ws.iter_cols(min_col=1, max_col=maxCol, min_row=1, max_row=6):
        for formCell in formCol:
            formCell.fill = whiteFill
    # # puts borders back on the cells
    for borderCol in ws.iter_cols(min_col=1, max_col=maxCol, min_row=6, max_row=maxRow):
        for borderCell in borderCol:
            borderCell.fill = no_fill

    # merges division with section info then merges cells if their not empty
    for sectionTitCol2 in ws.iter_cols(min_row=7, max_col=2, max_row=ws.max_row-10):
        for secCell2 in sectionTitCol2:
            if secCell2.value is not None:
                if secCell2.value == 'Division':
                    secCell2Row = secCell2.row
                    secCell2Value = secCell2.value
                    secCell2Neighbor = ws.cell(row=secCell2Row, column=2).value
                    secCell2.value = secCell2Value + ' ' + secCell2Neighbor
                ws.merge_cells(start_row=secCell2.row, start_column=1, end_row=secCell2.row, end_column=3)
                secCell2.fill = no_fill
                secCell2.border = no_border
            else:
                continue
    #  Colors Costs column blue and adds thin border
    for costCol in ws.iter_cols(min_row=7, max_row=ws.max_row, min_col=7, max_col=7):
        for costCell in costCol:
            if costCell.value is not None:
                costCell.fill = costFill
                costCell.border = thin_border
            else:
                continue
    # Step 1 finds the Finish section and gets
    for sectionTitCol in ws.iter_cols(min_row=1, max_col=2, max_row=ws.max_row):
        for sectionCell in sectionTitCol:  # row value
            if sectionCell.value == 'Division 18 - Finish':
                finStart = sectionCell.row
                # gets the row value and starts looking for Subtotal row
                for subTotCol in ws.iter_cols(min_col=3, min_row=finStart, max_col=3):
                    for subTotCell in subTotCol:  # subtotals in column C
                        if subTotCell.value == 'Sub Totals:':  # gets subtotals Row and colors Balance Cell
                            finSubTotRow = subTotCell.row
                            ws.cell(row=finSubTotRow, column=9).fill = totalsFill
                            ws.cell(row=finSubTotRow, column=9).border = thin_border
                            break
    # looks for grand totals and colors balance cell
    for grandTotalCol in ws.iter_cols(min_col=3, max_col=3):
        for grandCell in grandTotalCol:
            if grandCell.value == 'Grand Totals:':
                grandTotRow = grandCell.row
                ws.cell(row=grandTotRow, column=9).fill = totalsFill
                ws.cell(row=grandTotRow, column=9).border = thin_border
                break

    today = datetime.now().date()  # Gets today's date
    start = today - timedelta(days=today.weekday())  # Finds monday of current week for start of week format
    wb.save(start.strftime('%m-%d-%Y') + "-" + str(reportName) + ".xlsx")  # names file with start of week date in front
    #   of job cell info pulled from B4 plus .xlsx
